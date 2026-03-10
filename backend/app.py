from flask import Flask, request, jsonify
from datetime import datetime
from flask_cors import CORS
from core import BaikeRPABot
from models.db import Database
import threading

app = Flask(__name__)
CORS(app)

def format_task(task):
    """把 task 字典里的 datetime 转成字符串，避免 JSON 序列化失败"""
    if task and task.get('created_at'):
        task['created_at'] = task['created_at'].strftime('%Y-%m-%d %H:%M:%S')
    return task

db = Database()
db.init_tables()



# ── 健康检查（前端用来检测后端是否在线）────────────────────
@app.route('/api/ping', methods=['GET'])
def ping():
    return jsonify({'status': 'ok'})

# ── 单次搜索 ──────────────────────────────────────────────
@app.route('/api/search', methods=['POST'])
def search():
    data = request.get_json()
    keyword = data.get('keyword', '').strip()
    # mode: 'strict'（严格匹配）或 'related'（相关词条），默认 strict
    mode = data.get('mode', 'strict')
    show_browser = data.get('show_browser', False)

    if not keyword:
        return jsonify({'success': False, 'error': '关键词不能为空'}), 400

    print(f"收到搜索请求：{keyword}，模式：{mode}，显示浏览器：{show_browser}")

    # 创建任务（相关模式先用 1 个占位，抓完再补）
    task_id = db.create_task(task_name=f"单次搜索 - {keyword}", keywords=[keyword])
    print(f"创建任务成功，任务ID：{task_id}")

    items = db.get_task_items(task_id)
    if not items:
        return jsonify({'success': False, 'error': '创建任务失败'}), 500

    item_id = items[0]['id']
    db.update_task_progress(task_id, 1, keyword)

    bot = BaikeRPABot(headless=not show_browser, debug=True)

    try:
        if mode == 'related':
            # ── 相关模式 ──────────────────────────────────────
            # 1. 先用 RPA 拿到最多5条词条链接并逐一抓取
            results = bot.search_related(keyword)
            print(f"相关模式抓取完成，共 {len(results)} 条")

            # 2. 知道实际数量后，更新任务的 total_count
            #    并为每条结果补插 task_item（第一条已经在创建任务时插好了）
            db.update_task_total(task_id, len(results))

            for i, r in enumerate(results):
                if i == 0:
                    # 第一条直接更新已有的 item
                    db.update_item_result(item_id, r)
                else:
                    # 后续几条追加到同一个 task 下
                    new_item_id = db.add_task_item(task_id, r.get('title') or keyword)
                    db.update_item_result(new_item_id, r)

            db.update_task_counts(task_id)
            return jsonify({'success': True, 'mode': 'related', 'results': results})

        else:
            # 严格模式：返回单条（原有逻辑不变）
            result = bot.search_single(keyword)
            print(f"抓取完成：{result}")
            db.update_item_result(item_id, result)
            db.update_task_counts(task_id)

            # 无论成功失败都返回 200，让前端根据 success 字段判断
            return jsonify(result)

    except Exception as e:
        print(f"抓取过程出错：{str(e)}")
        error_result = {'success': False, 'keyword': keyword, 'error': str(e)}
        db.update_item_result(item_id, error_result)
        db.update_task_counts(task_id)
        return jsonify(error_result)   # 返回 200，前端靠 success:false 判断

    finally:
        bot.close()
        print("浏览器已关闭")


# ── 任务列表 ──────────────────────────────────────────────
@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    tasks = db.get_all_tasks()
    for task in tasks:
        total = task['total_count']
        task['progress'] = int(task['completed_count'] / total * 100) if total > 0 else 0
        format_task(task)
    return jsonify(tasks)


# ── 任务详情 ──────────────────────────────────────────────
@app.route('/api/tasks/<int:task_id>', methods=['GET'])
def get_task(task_id):
    task = db.get_task(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    total = task['total_count']
    task['progress'] = int(task['completed_count'] / total * 100) if total > 0 else 0
    return jsonify(format_task(task))


# ── 任务词条列表 ──────────────────────────────────────────
@app.route('/api/tasks/<int:task_id>/items', methods=['GET'])
def get_task_items(task_id):
    status = request.args.get('status')
    items = db.get_task_items(task_id, status)
    for item in items:
        if item.get('created_at'):
            item['created_at'] = item['created_at'].strftime('%Y-%m-%d %H:%M:%S')
    return jsonify(items)


# ── 删除任务 ──────────────────────────────────────────────
@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    success = db.delete_task(task_id)
    if not success:
        return jsonify({'error': '任务不存在'}), 404
    return jsonify({'success': True})



# ── 控制台统计 ────────────────────────────────────────────
@app.route('/api/stats', methods=['GET'])
def get_stats():
    stats = db.get_dashboard_stats()
    return jsonify(stats)

# ── 批量搜索：创建任务 ────────────────────────────────────
@app.route('/api/batch', methods=['POST'])
def batch_create():
    """
    接收关键词列表，创建任务并批量插入 task_items，
    然后在后台逐个执行搜索，每完成一条就更新数据库。
    返回 task_id，前端轮询 /api/tasks/<task_id> 获取进度。
    """
    data = request.get_json()
    keywords = data.get('keywords', [])
    keywords = [k.strip() for k in keywords if k.strip()]

    if not keywords:
        return jsonify({'success': False, 'error': '关键词列表不能为空'}), 400

    # 1. 创建任务，批量插入所有 item（状态全为 pending）
    task_id = db.create_task(
        task_name=f"批量搜索 - {len(keywords)} 个词条",
        keywords=keywords
    )
    print(f"批量任务创建成功，task_id={task_id}，共 {len(keywords)} 个词条")

    # 2. 后台线程执行，立即返回 task_id 让前端轮询
    def run_batch(task_id):
        # 后台线程用独立的数据库连接，避免与主线程共享连接导致断连
        from models.db import Database
        thread_db = Database()
        items = thread_db.get_task_items(task_id)
        bot   = BaikeRPABot(headless=True, debug=True)
        try:
            for i, item in enumerate(items):
                kw      = item['keyword']
                item_id = item['id']
                thread_db.update_task_progress(task_id, i + 1, kw)
                print(f"[{i+1}/{len(items)}] 开始抓取: {kw}")
                result = bot.search_single(kw)
                thread_db.update_item_result(item_id, result)
                thread_db.update_task_counts(task_id)
                print(f"[{i+1}/{len(items)}] 完成: {kw} -> {'成功' if result['success'] else '失败'}")
        except Exception as e:
            print(f"批量任务出错: {e}")
        finally:
            bot.close()
            try:
                thread_db.connection.close()
            except Exception:
                pass
            print(f"批量任务 {task_id} 执行完毕")

    thread = threading.Thread(target=run_batch, args=(task_id,), daemon=True)
    thread.start()

    return jsonify({'success': True, 'task_id': task_id})


# ── 来源占比统计 ──────────────────────────────────────────
@app.route('/api/stats/source', methods=['GET'])
def get_source_stats():
    """返回各来源（百度百科/墨鱼词典/AI）的词条数量"""
    data = db.get_source_stats()
    return jsonify(data)

# ── 近7天每日检索量 ───────────────────────────────────────
@app.route('/api/stats/weekly', methods=['GET'])
def get_weekly_stats():
    """返回近7天每天的检索词条数（成功+失败分开）"""
    stats = db.get_weekly_stats()
    return jsonify(stats)

# ── 导出任务结果为 Excel ──────────────────────────────────
@app.route('/api/tasks/<int:task_id>/export', methods=['GET'])
def export_task(task_id):
    """将任务词条结果导出为 Excel 文件"""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'error': '服务器缺少 openpyxl 依赖，请执行 pip install openpyxl'}), 500

    task = db.get_task(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    items = db.get_task_items(task_id)

    wb = Workbook()
    ws = wb.active
    ws.title = '检索结果'

    # ── 样式定义 ──
    header_font    = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill    = PatternFill('solid', fgColor='6366F1')
    center_align   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_border    = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )
    success_fill = PatternFill('solid', fgColor='D1FAE5')
    fail_fill    = PatternFill('solid', fgColor='FEE2E2')
    pending_fill = PatternFill('solid', fgColor='F3F4F6')

    # ── 第一行：任务信息标题 ──
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    task_name  = task['task_name'] if task else f'任务 {task_id}'
    title_cell.value     = f'检索任务：{task_name}'
    title_cell.font      = Font(name='微软雅黑', bold=True, size=13, color='1F2937')
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    # ── 第二行：统计摘要 ──
    ws.merge_cells('A2:G2')
    total   = task['total_count']     if task else len(items)
    success = task['success_count']   if task else 0
    failed  = task['failed_count']    if task else 0
    rate    = round(success / total * 100) if total > 0 else 0
    ws['A2'].value     = f'共 {total} 个词条   成功 {success}   失败 {failed}   成功率 {rate}%   导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font      = Font(name='微软雅黑', size=10, color='6B7280')
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 20

    # ── 空行 ──
    ws.row_dimensions[3].height = 6

    # ── 第四行：表头 ──
    headers = ['序号', '关键词', '标题', '摘要', '来源', '状态', '链接']
    col_widths = [6, 18, 22, 50, 12, 8, 40]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[4].height = 22

    # ── 数据行 ──
    status_label = {'success': '成功', 'failed': '失败', 'pending': '待处理'}
    for idx, item in enumerate(items, start=1):
        row = idx + 4
        status = item.get('status', 'pending')
        row_fill = success_fill if status == 'success' else (fail_fill if status == 'failed' else pending_fill)

        values = [
            idx,
            item.get('keyword', ''),
            item.get('title', ''),
            item.get('summary', '') or item.get('error_msg', ''),
            item.get('source', '百度百科'),
            status_label.get(status, status),
            item.get('url', ''),
        ]
        aligns = [center_align, left_align, left_align, left_align, center_align, center_align, left_align]

        for col, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = row_fill
            cell.alignment = aln
            cell.border    = thin_border
            cell.font      = Font(name='微软雅黑', size=10)

        # 链接列单独处理（蓝色可点击）
        url = item.get('url', '')
        if url:
            link_cell = ws.cell(row=row, column=7)
            link_cell.hyperlink = url
            link_cell.font = Font(name='微软雅黑', size=10, color='6366F1', underline='single')

        ws.row_dimensions[row].height = 18

    # ── 冻结首行表头 ──
    ws.freeze_panes = 'A5'

    # ── 输出文件流 ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    filename = f'检索结果_{task_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)